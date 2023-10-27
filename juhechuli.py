import openpyxl as op
from tqdm import tqdm
import spider as msd
import datetime  # 导入datetime模块
import pandas as pd
from main import CitiesCode
from main import ProvinceCode
import numpy as np
import csv
import os
import win32com.client as win32


def paqu(start_date, end_date, searchKey, migrationType, o_level, d_level,p,path):
    """
    Args:
        start_date (str): ''
        end_date (str): ''
        searchKey (str): "地级市名称/省名称"
        migrationType (str): 'in'/'out'
        o_level (str): 出发地行政级别,'city/province'
        d_level (str): 到达地行政级别,'city/province'
        p:'省份'
    """
    if not os.path.exists(path):
        os.makedirs(path)
    else:
        pass
    city = pd.read_json(p+'.json')
    NameList = list(city[searchKey].unique())
    start_date = start_date
    end_date = end_date
    global dateRange
    dateRange = pd.date_range(start_date, end_date).strftime("%Y-%m-%d").tolist()
    try:
        dateRange.remove('2021-08-25')
    except:
        pass
    try:
        dateRange.remove('2021-07-08')
    except:
        pass
    # try:
    #     dateRange.remove('2021-11-26')
    # except:
    #     pass
    # dateRange.remove('2021-07-08')
    # dateRange.remove('2021-08-12')
    # dateRange.remove('2021-08-25')
    batchdata = []
    for i in tqdm(NameList):
        a = msd.GetBatchDateData(o_level, d_level, i, migrationType=migrationType,date=dateRange,cityCodePath=p+".json")
        a.to_csv(path+"\\"+i+".csv",encoding="utf-8_sig")

    #         batchdata.append(a)
    #         pbar.update(1)
    # writer = pd.ExcelWriter(start_date+'to'+end_date + o_level+'to'+d_level + migrationType+'_'+ p +'.xlsx')
    # global mincheng
    # mincheng = start_date+'to'+end_date + o_level+'to'+d_level + migrationType +'_'+ p +'.xlsx'
    # for j in tqdm(batchdata):
    #     j.to_excel(writer, j.iloc[0][3])
    # writer.save()


def replace_excel(file):
    file = mincheng
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbooks.Open(file)
    wb.SaveAs(file + 'x', FileFormat=51)  # FileFormat = 51 is for .xlsx extension
    wb.Close()  # FileFormat = 56 is for .xls extension
    excel.Application.Quit()


def jisuan(start,end,p,mincheng,bb,c):
    i = 0
    global mm
    if c=='out':
        mm = '出'
    elif c == 'in':
        mm = '入'
    datestart = datetime.datetime.strptime(start, '%Y-%m-%d')
    dateend = datetime.datetime.strptime(end, '%Y-%m-%d')
    date_list = []
    while datestart < dateend:
        datestart += datetime.timedelta(days=1)
        date_list.append(datestart.strftime('%Y-%m-%d'))
    try:
        date_list.remove('2021-08-25')
    except:
        pass
    try:
        date_list.remove('2021-07-08')
    except:
        pass
    # try:
    #     date_list.remove('2021-11-26')
    # except:
    #     pass
    city = pd.read_json(p+'.json')
    NameList = list(city['地级市名称'].unique())
    # data1:迁出 data:迁入
    for nn in NameList:
        data = pd.read_csv(r'全国迁徙趋势\{}省\{}\迁{}\{}迁{}趋势.csv'.format("广东",nn,mm,nn,mm))
        index = data[data.date == bb].index.tolist()[0]
        data = np.array(data)
        reader = pd.ExcelFile(mincheng)
        dfs = []
        oo = -1
        ll = []
        for sheet in reader.sheet_names:
            df = reader.parse(sheet_name=sheet)
            dfs.append(df)
        for b in date_list:
            oo += 1
            ii = dfs[i][b] * 0.01 * data[index + oo][1]
            ll.append(ii)
            with open(r'{}to{}_{}_{}_{}有效计算.csv'.format(start,end,p,nn,mm), 'w') as csvFile:
                writer = csv.writer(csvFile, lineterminator='\n')
                writer.writerows(ll)
        i += 1

    for nnn in NameList:
        df = pd.read_csv(r'{}to{}_{}_{}_{}有效计算.csv'.format(start,end,p,nnn,mm))
        data = df.values
        index1 = list(df.keys())
        data = list(map(list, zip(*data)))
        data = pd.DataFrame(data, index=index1)
        data.to_csv(r'{}to{}_{}_{}_{}有效计算_转置.csv'.format(start,end,p,nnn,mm),header=0)

def zhuanhuan(start,end,p,searchKey):
    city = pd.read_json(p+'.json')
    NameList = list(city[searchKey].unique())
    print("正在修改原Ecxel数据........")
    for nnnn in tqdm(NameList):
        df = pd.read_csv(r'{}to{}_{}_{}_{}有效计算_转置.csv'.format(start,end,p,nnnn,mm),header=None)
        data = df.values
        bg = op.load_workbook(mincheng)
        sheet = bg[nnnn]
        # cell = sheet.cell(row=2, column=6) # 获取单元格指定行和指定列的内容
        # cell.value = 223
        for i in range(1, data.shape[0] + 1):
            for j in range(1, data.shape[1] + 1):
                sheet.cell(1 + i, 6 + j, data[-1 + i][j - 1])
        bg.save(mincheng)


def OD():
    city = pd.read_json(p + '.json')
    NameList = list(city['地级市名称'].unique())
    for nnnn in tqdm(NameList):
        df = pd.read_csv(r'{}to{}_{}_{}_{}有效计算_转置.csv'.format(start, end, p, nnnn, mm), header=None)



def shanchu(mincheng):
    reader = pd.ExcelFile(mincheng)
    dfs = []
    for sheet in reader.sheet_names:
        df = reader.parse(sheet_name=sheet)
        # 按列条件删除行
        df = df[df.city_name.isin(["广州市","佛山市","肇庆市","深圳市","东莞市","惠州市","珠海市","中山市","江门市"])]
        dfs.append(df)
    writer = pd.ExcelWriter(mincheng)
    for p in dfs:
        p.to_excel(writer,p.iloc[0][4])
    writer.save()


# 用于爬取比例数据
if __name__ == '__main__':
    start = '2023-01-01'
    end = '2023-05-31'
    p = 'CityCode'
    path = "./ceshi1"
    migrationType = 'in'
    print("开始爬取。。。。。。")
    paqu(start, end, '地级市名称', migrationType, 'city', 'city',p,path)
    print("空值填0.。。。。。。。。。。。。")
    nn = []
    reader = os.listdir(path)
    for sheet in tqdm(reader):
        df = pd.read_csv(path + "/" + sheet)
        df = df.fillna(0)
        df.to_csv(path+"\\" + sheet, encoding="utf-8_sig")
    # shanchu(mincheng)
    # jisuan(start,end,p,mincheng,20220602,migrationType)#start 后一天
    # zhuanhuan(start, end, p,'地级市名称')
    # city = pd.read_json(r'E:\百度人口迁徙\福建.json')
    # NameList = list(city['地级市名称'].unique())
    # print('正在画图......')
    # for i in tqdm(NameList):
    #     huatu(mincheng,i,migrationType, start)
    # # # # # city_trend()
